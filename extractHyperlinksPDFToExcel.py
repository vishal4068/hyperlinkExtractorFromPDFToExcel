# import fitz  # PyMuPDF
# from openpyxl import Workbook
# from openpyxl.styles import Font, Alignment

# def extract_hyperlink_text(pdf_path):
#     doc = fitz.open(pdf_path)
#     hyperlinks_info = []

#     for page_num, page in enumerate(doc, start=1):
#         links = page.get_links()
#         for link in links:
#             if link['kind'] == fitz.LINK_URI:
#                 rect = fitz.Rect(link['from'])
#                 words = page.get_text("words")
#                 link_text = " ".join([w[4] for w in words if fitz.Rect(w[:4]).intersects(rect)])
#                 hyperlink_url = link['uri']
#                 hyperlinks_info.append((link_text.strip(), hyperlink_url, page_num))

#     doc.close()
#     return hyperlinks_info

# def save_hyperlinks_to_excel(hyperlinks_info, excel_path):
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Hyperlinks"
#     ws.append(["Hyperlink Text", "Hyperlink URL", "Page Number"])  # Adding the header
#     ws = wb.active  # Assuming 'wb' is your Workbook instance and 'ws' is the active worksheet
#     for cell in ws[1]:  # Accessing the first row directly with ws[1]
#         cell.font = Font(bold=True)
#         cell.alignment = Alignment(horizontal='center')

#     for row, info in enumerate(hyperlinks_info, start=2):  # Start from row 2 to account for the header
#         hyperlink_text, hyperlink_url, page_number = info
#         # Insert the hyperlink text and page number
#         ws[f'A{row}'] = hyperlink_text
#         ws[f'C{row}'] = page_number
#         ws[f'C{row}'].alignment = Alignment(horizontal='center')
#         # Create a clickable hyperlink for the URL
#         ws[f'B{row}'].hyperlink = hyperlink_url
#         ws[f'B{row}'].value = hyperlink_url  # You can customize this text
#         ws[f'B{row}'].style = "Hyperlink"  # Optional: Apply the hyperlink style

#     wb.save(excel_path)

# # Example usage
# pdf_path = "vishal.pdf"
# excel_path = "hyperlink_text.xlsx"
# hyperlinks_info = extract_hyperlink_text(pdf_path)
# save_hyperlinks_to_excel(hyperlinks_info, excel_path)


import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import os
def extract_hyperlink_text(pdf_path):
    doc = fitz.open(pdf_path)
    hyperlinks_info = []

    for page_num, page in enumerate(doc, start=1):
        links = page.get_links()
        for link in links:
            if link['kind'] == fitz.LINK_URI:
                original_rect = fitz.Rect(link['from'])
                adjusted_rect = fitz.Rect(original_rect.x0, original_rect.y0 + 2, original_rect.x1, original_rect.y1 - 2)
                
                words = page.get_text("words")
                link_words = [w for w in words if fitz.Rect(w[:4]).intersects(adjusted_rect)]
                
                link_text = " ".join([w[4] for w in link_words])
                hyperlink_url = link['uri']
                hyperlinks_info.append((link_text.strip(), hyperlink_url, page_num))

    doc.close()

    # New step: Merge adjacent links with the same URL
    merged_hyperlinks_info = []
    previous_url = ""
    combined_text = ""
    for link_text, hyperlink_url, page_num in hyperlinks_info:
        if hyperlink_url == previous_url:
            # Combine text for adjacent links with the same URL
            combined_text += " " + link_text
        else:
            if previous_url:
                # Save the previous link if it's not the first iteration
                merged_hyperlinks_info.append((combined_text.strip(), previous_url, page_num))
            combined_text = link_text
        previous_url = hyperlink_url
    # Don't forget to add the last link
    if combined_text:
        merged_hyperlinks_info.append((combined_text.strip(), previous_url, page_num))

    return merged_hyperlinks_info

def save_hyperlinks_to_excel(hyperlinks_info, excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Hyperlinks"
    ws.append(["Page Number", "Hyperlink Text", "Hyperlink URL"])  # Adjusted the header order

    for cell in ws[1]:  # Formatting the header row
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row, info in enumerate(hyperlinks_info, start=2):
        hyperlink_text, hyperlink_url, page_number = info
        # Adjust the order here: Page number, Hyperlink text, and then Hyperlink URL
        ws[f'A{row}'] = page_number
        ws[f'A{row}'].alignment = Alignment(horizontal='center')  # Center-align the page number
        ws[f'B{row}'] = hyperlink_text
        # Create a clickable hyperlink for the URL
        ws[f'C{row}'].hyperlink = hyperlink_url
        ws[f'C{row}'].value = hyperlink_url  # You can customize this text
        ws[f'C{row}'].style = "Hyperlink"  # Optional: Apply the hyperlink style

    wb.save(excel_path)


# Example usage
pdf_path = "example.pdf"
# Extract the base name without the extension and append '_Extracted-Hyperlinks.xlsx'
excel_file_name = os.path.splitext(os.path.basename(pdf_path))[0] + "_Extracted-Hyperlinks.xlsx"
# Assuming you want to save the Excel file in the same directory as the PDF
excel_path = os.path.join(os.path.dirname(pdf_path), excel_file_name)

hyperlinks_info = extract_hyperlink_text(pdf_path)
save_hyperlinks_to_excel(hyperlinks_info, excel_path)