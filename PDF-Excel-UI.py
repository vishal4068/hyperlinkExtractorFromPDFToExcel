# Python script for PDF hyperlink extraction
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import os

def extract_hyperlink_text(pdf_path):
    doc = fitz.open(pdf_path)
    hyperlinks_info = []

    for page_num, page in enumerate(doc, start=1):
        links = page.get_links()
        for link in links:
            if link['kind'] == fitz.LINK_URI:
                original_rect = fitz.Rect(link['from'])
                adjusted_rect = fitz.Rect(original_rect.x0, original_rect.y0 + 2, original_rect.x1, original_rect.y1 - 2)
                
                words = page.get_text("words")
                link_words = [w for w in words if fitz.Rect(w[:4]).intersects(adjusted_rect)]
                
                link_text = " ".join([w[4] for w in link_words])
                hyperlink_url = link['uri']
                hyperlinks_info.append((link_text.strip(), hyperlink_url, page_num))

    doc.close()

    # New step: Merge adjacent links with the same URL
    merged_hyperlinks_info = []
    previous_url = ""
    combined_text = ""
    for link_text, hyperlink_url, page_num in hyperlinks_info:
        if hyperlink_url == previous_url:
            # Combine text for adjacent links with the same URL
            combined_text += " " + link_text
        else:
            if previous_url:
                # Save the previous link if it's not the first iteration
                merged_hyperlinks_info.append((combined_text.strip(), previous_url, page_num))
            combined_text = link_text
        previous_url = hyperlink_url
    # Don't forget to add the last link
    if combined_text:
        merged_hyperlinks_info.append((combined_text.strip(), previous_url, page_num))

    return merged_hyperlinks_info

def save_hyperlinks_to_excel(hyperlinks_info, excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Hyperlinks"
    ws.append(["Page Number", "Hyperlink Text", "Hyperlink URL"])  # Adjusted the header order

    for cell in ws[1]:  # Formatting the header row
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row, info in enumerate(hyperlinks_info, start=2):
        hyperlink_text, hyperlink_url, page_number = info
        # Adjust the order here: Page number, Hyperlink text, and then Hyperlink URL
        ws[f'A{row}'] = page_number
        ws[f'A{row}'].alignment = Alignment(horizontal='center')  # Center-align the page number
        ws[f'B{row}'] = hyperlink_text
        # Create a clickable hyperlink for the URL
        ws[f'C{row}'].hyperlink = hyperlink_url
        ws[f'C{row}'].value = hyperlink_url  # You can customize this text
        ws[f'C{row}'].style = "Hyperlink"  # Optional: Apply the hyperlink style

    wb.save(excel_path)

# Tkinter GUI for interacting with the PDF hyperlink extraction script
import tkinter as tk
from tkinter import filedialog, messagebox

def browse_pdf_file():
    filename = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
    if filename:
        pdf_entry.delete(0, tk.END)
        pdf_entry.insert(0, filename)

def extract_hyperlinks():
    pdf_path = pdf_entry.get()
    if not pdf_path:
        messagebox.showerror("Error", "Please select a PDF file.")
        return

    try:
        hyperlinks_info = extract_hyperlink_text(pdf_path)
        # Extract the base name of the PDF file without the extension
        base_name = os.path.splitext(os.path.basename(pdf_path))[0]
        # Construct the output Excel file name by appending '_Extracted-Hyperlinks.xlsx'
        excel_file_name = f"{base_name}_Extracted-Hyperlinks.xlsx"
        # Join the new file name with the directory of the input PDF to form the full path
        excel_path = os.path.join(os.path.dirname(pdf_path), excel_file_name)
        save_hyperlinks_to_excel(hyperlinks_info, excel_path)
        messagebox.showinfo("Success", f"Hyperlinks extracted successfully. Excel file saved as '{excel_path}'.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

# Create the main window
root = tk.Tk()
root.title("PDF Hyperlink Extractor")

# Create and place widgets
pdf_label = tk.Label(root, text="PDF File:")
pdf_label.grid(row=0, column=0, padx=10, pady=5)

pdf_entry = tk.Entry(root, width=50)
pdf_entry.grid(row=0, column=1, padx=10, pady=5)

browse_button = tk.Button(root, text="Browse", command=browse_pdf_file)
browse_button.grid(row=0, column=2, padx=10, pady=5)

extract_button = tk.Button(root, text="Extract Hyperlinks", command=extract_hyperlinks)

extract_button.grid(row=1, column=0, columnspan=3, pady=10)

# Run the Tkinter event loop
root.mainloop()
